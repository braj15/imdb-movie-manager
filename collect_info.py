import os
import json
import progressbar
import urllib.request
from imdbpie import Imdb
from guessit import guessit
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_movie_names(adir="."):

    exten = ('.mp4', '.avi', '.mkv', '.mpg', '.mov', '.vob', '.3gp', '.m2ts',
             '.3g2', '.flv', '.h264', '.mpeg', '.m4v', '.rm', '.wmv', '.swf')

    # Yields valid movie names
    for dirpath, dirnames, files in os.walk(adir):
        for name in files:
            if name.lower().endswith(exten) and not name.lower().startswith("sample"):
                yield ((guessit(name)).get('title'), dirpath)


def fetch_movie_info(name):

    # API key
    apikey = "78d08b59"

    try:
        # Gets movie info
        imdb = Imdb()
        imdb_search = imdb.search_for_title(name)
        movie_id = imdb_search[0]["imdb_id"]

        url = "http://www.omdbapi.com/?i=" + movie_id + "&apikey=" + apikey
        response = urllib.request.urlopen(url).read()
        jsonvalues = json.loads(response)

        if jsonvalues['Response'] == 'True':
            title = jsonvalues["Title"]
            rating = jsonvalues["imdbRating"]
            genre = jsonvalues["Genre"]
            year = jsonvalues["Year"]
            actors = jsonvalues["Actors"]
            director = jsonvalues["Director"]
            runtime = jsonvalues["Runtime"]

            # Gathers movie data in a list
            info_list = [movie_id, title, float(rating), genre, int(year),
                         actors, director, runtime]

            return (True, info_list)

        else:
            return (False, name)

    except Exception:
        return (False, name)


def create_collection():

    # Deletes existed .xlsx file
    if os.path.isfile("movie_info.xlsx"):
        os.remove("movie_info.xlsx")

    # Creates .xlsx file
    wb = Workbook()
    ws = wb.active

    # Adds Column headings
    ws.append(['Movie id', 'Movie name', 'Imdb rating', 'Genre',
               'Year', 'Actors', 'Director', 'Running time'])

    # Progressbar to see the progress
    bar = progressbar.ProgressBar(max_value=progressbar.UnknownLength)

    print()
    print("Reading your local directory for movies.\n")
    print("Fetching movie info....... Please wait!\n")

    end_row_no = 1

    # Set to add unique values only to the movie_names collection
    movie_names = set()

    # writes data
    for name, dirpath in get_movie_names():

        if name not in movie_names:
            abool, info_list = fetch_movie_info(name)

            if abool is True:
                movie_names.add(name)
                length = len(movie_names)
                bar.update(length)
                end_row_no += 1

                # Appends row
                ws.append(info_list)

                # Hyperlinks movie names
                ws.cell(row=end_row_no, column=2).value = '=HYPERLINK("{}", "{}")'.format(
                    dirpath, name)

    bar.finish()

    print()
    print(f"Found : {str(length)} movies in your collection.\n")
    print("Done!!!\n")

# Excel Design

    # Freezes the first row
    ws.freeze_panes = 'A2'

    # Adjusts column width based on its cell's content
    # excluding column two (because of hyperlink formulas)
    skip_col_no = 2
    for column_cells in ws.columns:
        if skip_col_no >= 0:
            skip_col_no -= 1
            if skip_col_no == 0:
                continue

        # Adjusts column width acc. value of max length (col2 excluded)
        value_length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column].width = value_length + 6

        # Center-aligns contents of the cells (col2 excluded)
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='center')

    # Adjusts 2nd col
    col2_val_length = max(len(nam) for nam in movie_names)
    ws.column_dimensions['B'].width = col2_val_length

    # Center-aligns col2
    ws['B1'].alignment = Alignment(horizontal='center')

    # Table
    tab = Table(displayName="movie_info", ref="A1:H"+str(end_row_no))

    # Adds a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)

    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save("movie_info.xlsx")


if __name__ == "__main__":
    create_collection()
